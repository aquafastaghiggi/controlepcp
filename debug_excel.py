import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('c:\\dadosCodi\\relatorio_api_2026.xlsx', data_only=True)
ws = wb.active

# Mostrar header
print("HEADER (colunas):")
for idx, cell in enumerate(ws[1]):
    if idx < 35:  # Mostrar primeiras 35 colunas
        print(f"  [{idx}] {cell.value}")

# Mostrar primeiras 5 linhas com datas
print("\n\nPRIMEIRAS 3 LINHAS (para diagnóstico de data):")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
    print(f"\nLinha {row_idx+2}:")
    for i in [13, 24, 30]:
        val = row[i] if len(row) > i else None
        print(f"  [{i}]: {val} (tipo: {type(val).__name__})")
