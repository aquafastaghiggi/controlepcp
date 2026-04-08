#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
import mysql.connector
from datetime import datetime
import json

print("[*] Iniciando importação de Excel para MySQL...")

# Conexão MySQL
try:
    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='k7m2y9u4',
        database='controlepcp_sandbox'
    )
    cursor = conn.cursor()
    print("[✓] Conectado ao MySQL")
except Exception as e:
    print(f"[✗] ERRO ao conectar MySQL: {e}")
    exit(1)

# Criar tabela se não existir
try:
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS relizado_2026_excel (
            id INT AUTO_INCREMENT PRIMARY KEY,
            data_evento DATE,
            ordem_op VARCHAR(20),
            quantidade DECIMAL(10, 2),
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uk_op_data (ordem_op, data_evento),
            INDEX idx_ordem_op (ordem_op),
            INDEX idx_data (data_evento)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    """)
    conn.commit()
    print("[✓] Tabela criada/verificada: relizado_2026_excel")
except Exception as e:
    print(f"[✗] ERRO ao criar tabela: {e}")
    exit(1)

# Limpar tabela antes de reimportar
try:
    cursor.execute("TRUNCATE TABLE relizado_2026_excel")
    conn.commit()
    print("[✓] Tabela limpa para reimportação")
except Exception as e:
    print(f"[✗] ERRO ao limpar tabela: {e}")

# Carregar Excel
try:
    wb = openpyxl.load_workbook('c:\\dadosCodi\\relatorio_api_2026.xlsx', data_only=True)
    ws = wb.active
    print(f"[✓] Excel carregado: {ws.max_row} linhas")
except Exception as e:
    print(f"[✗] ERRO ao carregar Excel: {e}")
    exit(1)

# Processar linhas
rows_processed = 0
rows_filtered = 0
rows_inserted = 0
rows_error = 0

for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
    rows_processed += 1
    
    try:
        # Extrair colunas (0-indexed)
        data_raw = row[1] if len(row) > 1 else None  # coluna 1 (B)
        ordem_op_raw = row[39] if len(row) > 39 else None  # coluna 39
        quantidade_raw = row[43] if len(row) > 43 else None  # coluna 43
        
        # Validar data
        if not data_raw:
            continue
        
        if not isinstance(data_raw, datetime):
            try:
                data_raw = datetime.fromisoformat(str(data_raw).split()[0])
            except:
                continue
        
        # Filtrar Mar-Abr 2026
        if not (data_raw.year == 2026 and data_raw.month in [3, 4]):
            continue
        
        rows_filtered += 1
        
        # Normalizar OP e quantidade
        if ordem_op_raw and quantidade_raw:
            ordem_op = str(ordem_op_raw).strip().lstrip('0') or '0'
            quantidade = float(quantidade_raw) if isinstance(quantidade_raw, (int, float)) else 0
            
            if ordem_op and quantidade > 0:
                try:
                    cursor.execute("""
                        INSERT INTO relizado_2026_excel (data_evento, ordem_op, quantidade)
                        VALUES (%s, %s, %s)
                        ON DUPLICATE KEY UPDATE quantidade = quantidade + VALUES(quantidade)
                    """, (data_raw.date(), ordem_op, quantidade))
                    rows_inserted += 1
                except mysql.connector.errors.IntegrityError:
                    pass  # Duplicate, atualizado com ON DUPLICATE KEY
                except Exception as e:
                    rows_error += 1
                    print(f"[!] Erro na linha {idx}: {e}")
    except Exception as e:
        rows_error += 1
        continue
    
    if rows_processed % 5000 == 0:
        conn.commit()
        print(f"[...] Processadas {rows_processed} linhas...")

# Commit final
conn.commit()

# Estatísticas
cursor.execute("SELECT COUNT(*) as total, SUM(quantidade) as soma FROM relizado_2026_excel")
result = cursor.fetchone()
total_rows = result[0]
total_qty = float(result[1] or 0)

cursor.execute("SELECT COUNT(DISTINCT ordem_op) as ops FROM relizado_2026_excel")
total_ops = cursor.fetchone()[0]

cursor.close()
conn.close()

print(f"""
[✓] IMPORTAÇÃO CONCLUÍDA
  - Linhas processadas: {rows_processed}
  - Linhas filtradas (Mar-Abr 2026): {rows_filtered}
  - Linhas inseridas/atualizadas: {rows_inserted}
  - Erros: {rows_error}
  - Total na tabela: {total_rows} registros
  - Total de OPs: {total_ops}
  - Quantidade total: {total_qty:,.2f}
""")
